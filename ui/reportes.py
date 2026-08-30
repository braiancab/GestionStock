from collections import defaultdict
from pathlib import Path
import customtkinter as ctk
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from tkinter import messagebox


class VentanaReportes:

  def __init__(self, ventana_padre):
    self.ventana_padre = ventana_padre

    self.ruta_excel_ventas = (
        Path.home()
        / "AppData"
        / "Local"
        / "GestionStock"
        / "datos"
        / "ventas.xlsx"
    )

    # ==========================================
    # VENTANA
    # ==========================================
    self.ventana = ctk.CTkToplevel(ventana_padre)
    self.ventana.title("Reporte de Artículos Más Vendidos")
    self.ventana.state("zoomed")
    self.ventana.transient(ventana_padre)
    self.ventana.grab_set()

    self.crear_interfaz()
    self.generar_reporte()

  def crear_interfaz(self):
    self.ventana.grid_columnconfigure(0, weight=1)
    self.ventana.grid_rowconfigure(1, weight=1)

    # Encabezado
    frame_top = ctk.CTkFrame(self.ventana, fg_color="transparent")
    frame_top.grid(row=0, column=0, padx=25, pady=(20, 10), sticky="ew")

    titulo = ctk.CTkLabel(
        frame_top,
        text="REPORTE DE ARTÍCULOS MÁS VENDEDOS",
        font=ctk.CTkFont(size=24, weight="bold"),
    )
    titulo.pack(side="left")

    btn_cerrar = ctk.CTkButton(
        frame_top,
        text="Cerrar",
        command=self.ventana.destroy,
        width=120,
        height=35,
    )
    btn_cerrar.pack(side="right")

    # Contenedor principal (Gráfico a la izquierda, Métrica / Resumen a la derecha)
    self.frame_contenido = ctk.CTkFrame(self.ventana)
    self.frame_contenido.grid(
        row=1, column=0, padx=25, pady=10, sticky="nsew"
    )
    self.frame_contenido.grid_columnconfigure(0, weight=3)
    self.frame_contenido.grid_columnconfigure(1, weight=1)
    self.frame_contenido.grid_rowconfigure(0, weight=1)

    # Frame para el gráfico Matplotlib
    self.frame_grafico = ctk.CTkFrame(self.frame_contenido)
    self.frame_grafico.grid(
        row=0, column=0, padx=15, pady=15, sticky="nsew"
    )

    # Frame para métricas laterales
    self.frame_metricas = ctk.CTkFrame(self.frame_contenido)
    self.frame_metricas.grid(
        row=0, column=1, padx=(0, 15), pady=15, sticky="nsew"
    )

  def cargar_datos_ventas(self):
    """Lee el archivo ventas.xlsx y consolida cantidades e ingresos por producto."""
    if not self.ruta_excel_ventas.exists():
      raise FileNotFoundError("No se encontró el archivo de ventas.xlsx.")

    wb = load_workbook(self.ruta_excel_ventas, data_only=True)
    ws = wb["Ventas"]

    cantidades = defaultdict(int)
    ingresos = defaultdict(float)
    total_recaudado = 0.0

    # Lectura de ítems vendidos
    for fila in ws.iter_rows(min_row=2, values_only=True):
      if not fila[0]:
        continue

      # Mapeo según la estructura de ventas.xlsx:
      # fila[4]: Producto | fila[5]: Variante | fila[6]: Cantidad | fila[8]: Subtotal
      prod_nombre = str(fila[4] or "Desconocido")
      variante = str(fila[5] or "").strip()

      if variante:
        etiqueta = f"{prod_nombre} ({variante})"
      else:
        etiqueta = prod_nombre

      cant = int(fila[6] or 0)
      subtotal = float(fila[8] or 0.0)

      cantidades[etiqueta] += cant
      ingresos[etiqueta] += subtotal
      total_recaudado += subtotal

    wb.close()

    # Ordenar productos por cantidad vendida (Top 10)
    top_productos = sorted(
        cantidades.items(), key=lambda item: item[1], reverse=True
    )[:10]

    return top_productos, ingresos, total_recaudado

  def generar_reporte(self):
    try:
      top_productos, ingresos, total_recaudado = (
          self.cargar_datos_ventas()
      )

      if not top_productos:
        messagebox.showwarning(
            "Reporte",
            "No hay registros de ventas para procesar.",
            parent=self.ventana,
        )
        return

      # Preparar datos para Matplotlib
      nombres = [item[0] for item in top_productos]
      cantidades = [item[1] for item in top_productos]

      # Estilo oscuro/adaptado del gráfico
      plt.style.use("dark_background")
      fig, ax = plt.subplots(figsize=(7, 5), dpi=100)
      fig.patch.set_facecolor("#2b2b2b")
      ax.set_facecolor("#2b2b2b")

      # Dibujar barras verticales
      barras = ax.bar(
          nombres, cantidades, color="#1f538d", edgecolor="white", alpha=0.85
      )

      ax.set_ylabel("Unidades Vendidas", fontsize=11, color="white")
      ax.set_title(
          "Top 10 Productos Más Vendidos",
          fontsize=14,
          fontweight="bold",
          color="white",
          pad=15,
      )

      # Añadir valores numéricos encima de cada barra
      for barra in barras:
        alto = barra.get_height()
        ax.annotate(
            f"{int(alto)}",
            xy=(barra.get_x() + barra.get_width() / 2, alto),
            xytext=(0, 3),
            textcoords="offset points",
            ha="center",
            va="bottom",
            color="white",
            fontsize=10,
            fontweight="bold",
        )

      plt.xticks(rotation=25, ha="right", fontsize=9, color="white")
      plt.yticks(color="white")
      ax.spines["top"].set_visible(False)
      ax.spines["right"].set_visible(False)
      plt.tight_layout()

      # Renderizar gráfico en Tkinter
      canvas = FigureCanvasTkAgg(fig, master=self.frame_grafico)
      canvas.draw()
      canvas.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=10)

      # Renderizar panel de métricas laterales
      self.mostrar_metricas(
          top_productos, ingresos, total_recaudado
      )

    except FileNotFoundError as err:
      messagebox.showwarning("Atención", str(err), parent=self.ventana)
    except Exception as err:
      messagebox.showerror(
          "Error",
          f"No se pudo generar el reporte:\n\n{err}",
          parent=self.ventana,
      )

  def mostrar_metricas(self, top_productos, ingresos, total_recaudado):
    # Título métricas
    lbl_tit = ctk.CTkLabel(
        self.frame_metricas,
        text="RESUMEN FINANCIERO",
        font=ctk.CTkFont(size=18, weight="bold"),
    )
    lbl_tit.pack(pady=(20, 15))

    # Caja de Total General
    card_total = ctk.CTkFrame(self.frame_metricas, fg_color="#1f538d")
    card_total.pack(fill="x", padx=15, pady=10)

    ctk.CTkLabel(
        card_total,
        text="RECAUDACIÓN TOTAL",
        font=ctk.CTkFont(size=12, weight="bold"),
        text_color="white",
    ).pack(pady=(10, 2))
    ctk.CTkLabel(
        card_total,
        text=f"${total_recaudado:,.2f}",
        font=ctk.CTkFont(size=22, weight="bold"),
        text_color="white",
    ).pack(pady=(0, 10))

    # Detalle de ingresos por producto Top
    lbl_det = ctk.CTkLabel(
        self.frame_metricas,
        text="Ingresos por Producto (Top):",
        font=ctk.CTkFont(size=13, weight="bold"),
    )
    lbl_det.pack(anchor="w", padx=15, pady=(15, 5))

    frame_scroll = ctk.CTkScrollableFrame(self.frame_metricas)
    frame_scroll.pack(fill="both", expand=True, padx=15, pady=(0, 15))

    for prod, cant in top_productos:
      monto = ingresos[prod]
      card_p = ctk.CTkFrame(frame_scroll, fg_color="transparent")
      card_p.pack(fill="x", pady=4)

      ctk.CTkLabel(
          card_p,
          text=f"• {prod}",
          font=ctk.CTkFont(size=12, weight="bold"),
          anchor="w",
      ).pack(fill="x")
      ctk.CTkLabel(
          card_p,
          text=f"  {cant} unid. | ${monto:,.2f}",
          font=ctk.CTkFont(size=11),
          text_color="gray",
          anchor="w",
      ).pack(fill="x")