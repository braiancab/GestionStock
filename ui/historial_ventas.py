import customtkinter as ctk
from tkinter import ttk, messagebox

from pathlib import Path
from openpyxl import load_workbook

class HistorialVentas:

    def __init__(self, ventana_padre):

        self.ventana_padre = ventana_padre

        self.ruta_excel = (
            Path.home()
            / "AppData"
            / "Local"
            / "GestionStock"
            / "datos"
            / "ventas.xlsx"
        )

        self.ventas = []

        # ==========================================
        # VENTANA
        # ==========================================

        self.ventana = ctk.CTkToplevel(
            ventana_padre
        )

        self.ventana.title(
            "Historial de ventas"
        )

        self.ventana.state("zoomed")

        self.ventana.transient(
            ventana_padre
        )

        self.ventana.grab_set()

        self.crear_interfaz()

        self.cargar_ventas()


    # ==========================================
    # INTERFAZ
    # ==========================================

    def crear_interfaz(self):

        self.ventana.grid_columnconfigure(
            0,
            weight=1
        )

        self.ventana.grid_rowconfigure(
            2,
            weight=1
        )


        # ==========================================
        # TÍTULO
        # ==========================================

        titulo = ctk.CTkLabel(
            self.ventana,
            text="HISTORIAL DE VENTAS",
            font=ctk.CTkFont(
                size=26,
                weight="bold"
            )
        )

        titulo.grid(
            row=0,
            column=0,
            padx=25,
            pady=(20, 10),
            sticky="w"
        )


        # ==========================================
        # FILTROS
        # ==========================================

        frame_filtros = ctk.CTkFrame(
            self.ventana
        )

        frame_filtros.grid(
            row=1,
            column=0,
            padx=25,
            pady=10,
            sticky="ew"
        )


        label_cliente = ctk.CTkLabel(
            frame_filtros,
            text="Buscar cliente:"
        )

        label_cliente.grid(
            row=0,
            column=0,
            padx=10,
            pady=15
        )


        self.entry_cliente = ctk.CTkEntry(
            frame_filtros,
            width=250,
            placeholder_text="Nombre del cliente"
        )

        self.entry_cliente.grid(
            row=0,
            column=1,
            padx=10,
            pady=15
        )


        boton_buscar = ctk.CTkButton(
            frame_filtros,
            text="Buscar",
            command=self.buscar_ventas
        )

        boton_buscar.grid(
            row=0,
            column=2,
            padx=10
        )


        boton_todas = ctk.CTkButton(
            frame_filtros,
            text="Mostrar todas",
            command=self.mostrar_todas
        )

        boton_todas.grid(
            row=0,
            column=3,
            padx=10
        )


        # ==========================================
        # TABLA
        # ==========================================

        frame_tabla = ctk.CTkFrame(
            self.ventana
        )

        frame_tabla.grid(
            row=2,
            column=0,
            padx=25,
            pady=10,
            sticky="nsew"
        )

        frame_tabla.grid_rowconfigure(
            0,
            weight=1
        )

        frame_tabla.grid_columnconfigure(
            0,
            weight=1
        )


        columnas = (
            "id",
            "fecha",
            "cliente",
            "total"
        )


        self.tabla = ttk.Treeview(
            frame_tabla,
            columns=columnas,
            show="headings"
        )


        self.tabla.heading(
            "id",
            text="Venta"
        )

        self.tabla.heading(
            "fecha",
            text="Fecha"
        )

        self.tabla.heading(
            "cliente",
            text="Cliente"
        )

        self.tabla.heading(
            "total",
            text="Total"
        )


        self.tabla.column(
            "id",
            width=100,
            anchor="center"
        )

        self.tabla.column(
            "fecha",
            width=180
        )

        self.tabla.column(
            "cliente",
            width=300
        )

        self.tabla.column(
            "total",
            width=180,
            anchor="e"
        )


        self.tabla.grid(
            row=0,
            column=0,
            sticky="nsew"
        )


        scrollbar = ttk.Scrollbar(
            frame_tabla,
            orient="vertical",
            command=self.tabla.yview
        )

        scrollbar.grid(
            row=0,
            column=1,
            sticky="ns"
        )

        self.tabla.configure(
            yscrollcommand=scrollbar.set
        )


        # DOBLE CLICK

        self.tabla.bind(
            "<Double-1>",
            self.ver_detalle
        )


        # ==========================================
        # PARTE INFERIOR
        # ==========================================

        frame_inferior = ctk.CTkFrame(
            self.ventana,
            fg_color="transparent"
        )

        frame_inferior.grid(
            row=3,
            column=0,
            padx=25,
            pady=15,
            sticky="ew"
        )


        self.label_cantidad = ctk.CTkLabel(
            frame_inferior,
            text="Ventas: 0",
            font=ctk.CTkFont(
                size=16
            )
        )

        self.label_cantidad.grid(
            row=0,
            column=0,
            padx=10
        )


        self.label_total = ctk.CTkLabel(
            frame_inferior,
            text="TOTAL VENDIDO: $0.00",
            font=ctk.CTkFont(
                size=20,
                weight="bold"
            )
        )

        self.label_total.grid(
            row=0,
            column=1,
            padx=30
        )


        boton_detalle = ctk.CTkButton(
            frame_inferior,
            text="Ver detalle",
            command=self.ver_detalle
        )

        boton_detalle.grid(
            row=0,
            column=2,
            padx=10
        )


        boton_cerrar = ctk.CTkButton(
            frame_inferior,
            text="Cerrar",
            command=self.ventana.destroy
        )

        boton_cerrar.grid(
            row=0,
            column=3,
            padx=10
        )


    # ==========================================
    # CARGAR VENTAS
    # ==========================================

    def cargar_ventas(self):

        try:

            wb = load_workbook(
                self.ruta_excel,
                data_only=True
            )

            ws = wb["Ventas"]

            ventas_dict = {}


            for fila in ws.iter_rows(
                min_row=2,
                values_only=True
            ):

                if not fila[0]:
                    continue


                id_venta = fila[0]

                fecha = fila[1]

                cliente = fila[2]

                total = fila[8]


                if id_venta not in ventas_dict:

                    ventas_dict[id_venta] = {

                        "id": id_venta,

                        "fecha": fecha,

                        "cliente": cliente,

                        "total": total or 0

                    }


            wb.close()


            self.ventas = list(
                ventas_dict.values()
            )


            # Más recientes primero

            self.ventas.sort(
                key=lambda x: x["id"],
                reverse=True
            )


            self.mostrar_ventas(
                self.ventas
            )


        except FileNotFoundError:

            messagebox.showwarning(
                "Historial",
                "Todavía no existe el archivo de ventas.",
                parent=self.ventana
            )


        except Exception as error:

            messagebox.showerror(
                "Error",
                f"No se pudo cargar el historial.\n\n"
                f"{error}",
                parent=self.ventana
            )


    # ==========================================
    # MOSTRAR VENTAS
    # ==========================================

    def mostrar_ventas(
        self,
        ventas
    ):

        for item in self.tabla.get_children():

            self.tabla.delete(item)


        total_general = 0


        for venta in ventas:

            total = venta["total"] or 0

            total_general += total


            self.tabla.insert(

                "",

                "end",

                values=(

                    venta["id"],

                    venta["fecha"],

                    venta["cliente"],

                    f"${total:,.2f}"

                )

            )


        self.label_cantidad.configure(

            text=f"Ventas: {len(ventas)}"

        )


        self.label_total.configure(

            text=f"TOTAL VENDIDO: ${total_general:,.2f}"

        )


    # ==========================================
    # BUSCAR CLIENTE
    # ==========================================

    def buscar_ventas(self):

        texto = (
            self.entry_cliente.get()
            .strip()
            .lower()
        )


        if not texto:

            self.mostrar_ventas(
                self.ventas
            )

            return


        resultados = []


        for venta in self.ventas:

            cliente = str(
                venta["cliente"] or ""
            ).lower()


            if texto in cliente:

                resultados.append(
                    venta
                )


        self.mostrar_ventas(
            resultados
        )


    # ==========================================
    # MOSTRAR TODAS
    # ==========================================

    def mostrar_todas(self):

        self.entry_cliente.delete(
            0,
            "end"
        )

        self.mostrar_ventas(
            self.ventas
        )


    # ==========================================
    # VER DETALLE
    # ==========================================

    def ver_detalle(
        self,
        evento=None
    ):

        seleccion = (
            self.tabla.selection()
        )


        if not seleccion:

            messagebox.showwarning(
                "Detalle",
                "Seleccione una venta.",
                parent=self.ventana
            )

            return


        valores = self.tabla.item(
            seleccion[0],
            "values"
        )


        id_venta = int(
            valores[0]
        )


        self.mostrar_detalle(
            id_venta
        )


    # ==========================================
    # MOSTRAR DETALLE
    # ==========================================

    def mostrar_detalle(
        self,
        id_venta
    ):

        try:

            wb = load_workbook(
                self.ruta_excel,
                data_only=True
            )

            ws = wb["Ventas"]


            productos = []

            cliente = ""

            fecha = ""

            total_venta = 0


            for fila in ws.iter_rows(
                min_row=2,
                values_only=True
            ):

                if fila[0] != id_venta:

                    continue


                cliente = fila[2]

                fecha = fila[1]

                total_venta = fila[9] or 0


                productos.append({

                    "producto": fila[4],

                    "variante": fila[5],

                    "cantidad": fila[6],

                    "precio": fila[7],

                    "subtotal": fila[8]

                })


            wb.close()


            if not productos:

                messagebox.showwarning(
                    "Detalle",
                    "No se encontraron productos para esta venta.",
                    parent=self.ventana
                )

                return


            # ==========================================
            # VENTANA DETALLE
            # ==========================================

            ventana_detalle = ctk.CTkToplevel(
                self.ventana
            )

            ventana_detalle.title(
                f"Venta #{id_venta}"
            )

            ventana_detalle.geometry(
                "850x600"
            )

            ventana_detalle.minsize(
                750,
                500
            )

            ventana_detalle.transient(
                self.ventana
            )


            titulo = ctk.CTkLabel(

                ventana_detalle,

                text=f"DETALLE DE VENTA #{id_venta}",

                font=ctk.CTkFont(
                    size=24,
                    weight="bold"
                )

            )

            titulo.pack(
                pady=(20, 10)
            )


            info = ctk.CTkLabel(

                ventana_detalle,

                text=(
                    f"Cliente: {cliente}\n"
                    f"Fecha: {fecha}"
                ),

                font=ctk.CTkFont(
                    size=15
                )

            )

            info.pack(
                pady=10
            )


            # ==========================================
            # TABLA
            # ==========================================

            frame_tabla = ctk.CTkFrame(
                ventana_detalle
            )

            frame_tabla.pack(
                fill="both",
                expand=True,
                padx=20,
                pady=(10, 5)
            )


            columnas = (
                "producto",
                "variante",
                "cantidad",
                "precio",
                "subtotal"
            )


            tabla = ttk.Treeview(

                frame_tabla,

                columns=columnas,

                show="headings"

            )


            tabla.heading(
                "producto",
                text="Producto"
            )

            tabla.heading(
                "variante",
                text="Variante"
            )

            tabla.heading(
                "cantidad",
                text="Cantidad"
            )

            tabla.heading(
                "precio",
                text="Precio"
            )

            tabla.heading(
                "subtotal",
                text="Subtotal"
            )


            tabla.column(
                "producto",
                width=220
            )

            tabla.column(
                "variante",
                width=130
            )

            tabla.column(
                "cantidad",
                width=90,
                anchor="center"
            )

            tabla.column(
                "precio",
                width=120,
                anchor="e"
            )

            tabla.column(
                "subtotal",
                width=130,
                anchor="e"
            )


            tabla.pack(
                fill="both",
                expand=True
            )


            for producto in productos:

                tabla.insert(

                    "",

                    "end",

                    values=(

                        producto["producto"],

                        producto["variante"],

                        producto["cantidad"],

                        f"${producto['precio']:,.2f}",

                        f"${producto['subtotal']:,.2f}"

                    )

                )


            # ==========================================
            # PARTE INFERIOR
            # ==========================================

            frame_inferior = ctk.CTkFrame(
                ventana_detalle,
                fg_color="transparent"
            )

            frame_inferior.pack(
                fill="x",
                padx=20,
                pady=(5, 20)
            )


            total = ctk.CTkLabel(
                frame_inferior,
                text=f"TOTAL: ${total_venta:,.2f}",
                font=ctk.CTkFont(
                    size=22,
                    weight="bold"
                )
            )

            total.pack(
                pady=(5, 12)
            )


            boton_cerrar = ctk.CTkButton(
                frame_inferior,
                text="Cerrar",
                command=ventana_detalle.destroy,
                width=140,
                height=35
            )

            boton_cerrar.pack()

        except Exception as error:

            messagebox.showerror(

                "Error",

                f"No se pudo mostrar el detalle.\n\n"
                f"{error}",

                parent=self.ventana

            )