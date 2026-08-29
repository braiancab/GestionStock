import os
import subprocess
import sys
from pathlib import Path
import customtkinter as ctk
from openpyxl import load_workbook
from tkinter import ttk, messagebox

from datetime import datetime
class HistorialPresupuestos:

  def __init__(self, ventana_padre):
    self.ventana_padre = ventana_padre

    # Ruta predeterminada al archivo Excel de presupuestos
    self.ruta_excel = (
        Path.home()
        / "AppData"
        / "Local"
        / "GestionStock"
        / "datos"
        / "presupuestos.xlsx"
    )

    # Carpeta donde se guardan/descargan los PDFs de presupuestos
    self.carpeta_pdf = (
        Path.home()
        / "AppData"
        / "Local"
        / "GestionStock"
        / "presupuestos_pdf"
    )

    self.presupuestos = []

    # ==========================================
    # VENTANA
    # ==========================================
    self.ventana = ctk.CTkToplevel(ventana_padre)
    self.ventana.title("Historial de Presupuestos")
    self.ventana.state("zoomed")
    self.ventana.transient(ventana_padre)
    self.ventana.grab_set()

    self.crear_interfaz()
    self.cargar_presupuestos()

  # ==========================================
  # INTERFAZ
  # ==========================================
  def crear_interfaz(self):
    self.ventana.grid_columnconfigure(0, weight=1)
    self.ventana.grid_rowconfigure(2, weight=1)

    # TÍTULO
    titulo = ctk.CTkLabel(
        self.ventana,
        text="HISTORIAL DE PRESUPUESTOS",
        font=ctk.CTkFont(size=26, weight="bold"),
    )
    titulo.grid(row=0, column=0, padx=25, pady=(20, 10), sticky="w")

    # FILTROS
    frame_filtros = ctk.CTkFrame(self.ventana)
    frame_filtros.grid(row=1, column=0, padx=25, pady=10, sticky="ew")

    label_cliente = ctk.CTkLabel(frame_filtros, text="Buscar cliente:")
    label_cliente.grid(row=0, column=0, padx=10, pady=15)

    self.entry_cliente = ctk.CTkEntry(
        frame_filtros, width=250, placeholder_text="Nombre del cliente"
    )
    self.entry_cliente.grid(row=0, column=1, padx=10, pady=15)

    boton_buscar = ctk.CTkButton(
        frame_filtros, text="Buscar", command=self.buscar_presupuestos
    )
    boton_buscar.grid(row=0, column=2, padx=10)

    boton_todas = ctk.CTkButton(
        frame_filtros, text="Mostrar todos", command=self.mostrar_todos
    )
    boton_todas.grid(row=0, column=3, padx=10)

    # TABLA
    frame_tabla = ctk.CTkFrame(self.ventana)
    frame_tabla.grid(row=2, column=0, padx=25, pady=10, sticky="nsew")
    frame_tabla.grid_rowconfigure(0, weight=1)
    frame_tabla.grid_columnconfigure(0, weight=1)

    columnas = ("id", "fecha", "cliente", "total")

    self.tabla = ttk.Treeview(
        frame_tabla, columns=columnas, show="headings", selectmode="browse"
    )

    self.tabla.heading("id", text="Nº Presupuesto")
    self.tabla.heading("fecha", text="Fecha")
    self.tabla.heading("cliente", text="Cliente")
    self.tabla.heading("total", text="Total")

    self.tabla.column("id", width=120, anchor="center")
    self.tabla.column("fecha", width=180, anchor="center")
    self.tabla.column("cliente", width=300)
    self.tabla.column("total", width=180, anchor="e")

    self.tabla.grid(row=0, column=0, sticky="nsew")

    scrollbar = ttk.Scrollbar(
        frame_tabla, orient="vertical", command=self.tabla.yview
    )
    scrollbar.grid(row=0, column=1, sticky="ns")
    self.tabla.configure(yscrollcommand=scrollbar.set)

    self.tabla.bind("<Double-1>", self.ver_detalle)

    # PARTE INFERIOR
    frame_inferior = ctk.CTkFrame(self.ventana, fg_color="transparent")
    frame_inferior.grid(row=3, column=0, padx=25, pady=15, sticky="ew")

    self.label_cantidad = ctk.CTkLabel(
        frame_inferior, text="Presupuestos: 0", font=ctk.CTkFont(size=16)
    )
    self.label_cantidad.grid(row=0, column=0, padx=10)

    self.label_total = ctk.CTkLabel(
        frame_inferior,
        text="TOTAL PRESUPUESTADO: $0.00",
        font=ctk.CTkFont(size=20, weight="bold"),
    )
    self.label_total.grid(row=0, column=1, padx=30)

    boton_detalle = ctk.CTkButton(
        frame_inferior, text="Ver detalle", command=self.ver_detalle
    )
    boton_detalle.grid(row=0, column=2, padx=10)

    boton_cerrar = ctk.CTkButton(
        frame_inferior, text="Cerrar", command=self.ventana.destroy
    )
    boton_cerrar.grid(row=0, column=3, padx=10)

  # ==========================================
  # CARGAR PRESUPUESTOS DESDE EXCEL
  # ==========================================
  def cargar_presupuestos(self):
    try:
      wb = load_workbook(self.ruta_excel, data_only=True)
      ws = wb["Presupuestos"]

      presupuestos_dict = {}

      for fila in ws.iter_rows(min_row=2, values_only=True):
        if not fila[0]:
          continue

        id_presupuesto = fila[0]
        fecha = fila[1]
        cliente = fila[2]
        total = fila[8]  # Columna Total General en el Excel

        if id_presupuesto not in presupuestos_dict:
          presupuestos_dict[id_presupuesto] = {
              "id": id_presupuesto,
              "fecha": fecha,
              "cliente": cliente,
              "total": total or 0,
          }

      wb.close()

      #self.presupuestos = list(presupuestos_dict.values())
      #self.presupuestos.sort(key=lambda x: str(x["id"]), reverse=True)
      self.presupuestos = list(presupuestos_dict.values())

      # Función para parsear la fecha y hora almacenada (formato: "DD/MM/YYYY HH:MM:SS")
      def obtener_fecha_dt(item):
        try:
          return datetime.strptime(str(item["fecha"]), "%d/%m/%Y %H:%M:%S")
        except Exception:
          # Si falla el parseo o la fecha no coincide exactamente con el formato,
          # intenta usar el ID o una fecha mínima de respaldo
          return datetime.min


      # Ordenar de más reciente a más antiguo (reverse=True)
      self.presupuestos.sort(key=obtener_fecha_dt, reverse=True)



      self.mostrar_presupuestos(self.presupuestos)

    except FileNotFoundError:
      messagebox.showwarning(
          "Historial",
          "Todavía no existe el archivo de presupuestos.",
          parent=self.ventana,
      )

    except Exception as error:
      messagebox.showerror(
          "Error",
          f"No se pudo cargar el historial de presupuestos.\n\n{error}",
          parent=self.ventana,
      )

  # ==========================================
  # MOSTRAR PRESUPUESTOS EN TABLA
  # ==========================================
  def mostrar_presupuestos(self, presupuestos):
    for item in self.tabla.get_children():
      self.tabla.delete(item)

    total_general = 0

    for presupuesto in presupuestos:
      total = presupuesto["total"] or 0
      total_general += total

      self.tabla.insert(
          "",
          "end",
          values=(
              presupuesto["id"],
              presupuesto["fecha"],
              presupuesto["cliente"],
              f"${total:,.2f}",
          ),
      )

    self.label_cantidad.configure(
        text=f"Presupuestos: {len(presupuestos)}"
    )
    self.label_total.configure(
        text=f"TOTAL PRESUPUESTADO: ${total_general:,.2f}"
    )

  # ==========================================
  # BUSCAR / FILTRAR CLIENTE
  # ==========================================
  def buscar_presupuestos(self):
    texto = self.entry_cliente.get().strip().lower()

    if not texto:
      self.mostrar_presupuestos(self.presupuestos)
      return

    resultados = [
        p
        for p in self.presupuestos
        if texto in str(p["cliente"] or "").lower()
    ]
    self.mostrar_presupuestos(resultados)

  def mostrar_todos(self):
    self.entry_cliente.delete(0, "end")
    self.mostrar_presupuestos(self.presupuestos)

  # ==========================================
  # VER DETALLE
  # ==========================================
  def ver_detalle(self, evento=None):
    seleccion = self.tabla.selection()

    if not seleccion:
      messagebox.showwarning(
          "Detalle",
          "Seleccione un presupuesto.",
          parent=self.ventana,
      )
      return

    valores = self.tabla.item(seleccion[0], "values")
    id_presupuesto = valores[0]

    self.mostrar_detalle(id_presupuesto)

  # ==========================================
  # MOSTRAR DETALLE Y OPCIÓN PDF
  # ==========================================
  def mostrar_detalle(self, id_presupuesto):
    try:
      wb = load_workbook(self.ruta_excel, data_only=True)
      ws = wb["Presupuestos"]

      productos = []
      cliente = ""
      fecha = ""
      total_presupuesto = 0

      for fila in ws.iter_rows(min_row=2, values_only=True):
        if str(fila[0]) != str(id_presupuesto):
          continue

        cliente = fila[2]
        fecha = fila[1]
        total_presupuesto = fila[9] or 0  # Columna Total acumulado/final

        productos.append({
            "producto": fila[4],
            "variante": fila[5],
            "cantidad": fila[6],
            "precio": fila[7],
            "subtotal": fila[8],
        })

      wb.close()

      if not productos:
        messagebox.showwarning(
            "Detalle",
            "No se encontraron productos para este presupuesto.",
            parent=self.ventana,
        )
        return

      # VENTANA EMERGENTE DETALLE
      ventana_detalle = ctk.CTkToplevel(self.ventana)
      ventana_detalle.title(f"Presupuesto #{id_presupuesto}")
      ventana_detalle.geometry("850x620")
      ventana_detalle.minsize(750, 500)
      ventana_detalle.transient(self.ventana)
      ventana_detalle.grab_set()

      titulo = ctk.CTkLabel(
          ventana_detalle,
          text=f"DETALLE DE PRESUPUESTO #{id_presupuesto}",
          font=ctk.CTkFont(size=24, weight="bold"),
      )
      titulo.pack(pady=(20, 10))

      info = ctk.CTkLabel(
          ventana_detalle,
          text=f"Cliente: {cliente}\nFecha: {fecha}",
          font=ctk.CTkFont(size=15),
      )
      info.pack(pady=10)

      # TABLA DE PRODUCTOS
      frame_tabla = ctk.CTkFrame(ventana_detalle)
      frame_tabla.pack(
          fill="both", expand=True, padx=20, pady=(10, 5)
      )

      columnas = ("producto", "variante", "cantidad", "precio", "subtotal")
      tabla = ttk.Treeview(frame_tabla, columns=columnas, show="headings")

      tabla.heading("producto", text="Producto")
      tabla.heading("variante", text="Variante")
      tabla.heading("cantidad", text="Cantidad")
      tabla.heading("precio", text="Precio")
      tabla.heading("subtotal", text="Subtotal")

      tabla.column("producto", width=220)
      tabla.column("variante", width=130)
      tabla.column("cantidad", width=90, anchor="center")
      tabla.column("precio", width=120, anchor="e")
      tabla.column("subtotal", width=130, anchor="e")

      tabla.pack(fill="both", expand=True)

      for prod in productos:
        tabla.insert(
            "",
            "end",
            values=(
                prod["producto"],
                prod["variante"],
                prod["cantidad"],
                f"${prod['precio']:,.2f}",
                f"${prod['subtotal']:,.2f}",
            ),
        )

      # PARTE INFERIOR DETALLE
      frame_inferior = ctk.CTkFrame(ventana_detalle, fg_color="transparent")
      frame_inferior.pack(fill="x", padx=20, pady=(5, 20))

      total = ctk.CTkLabel(
          frame_inferior,
          text=f"TOTAL: ${total_presupuesto:,.2f}",
          font=ctk.CTkFont(size=22, weight="bold"),
      )
      total.pack(pady=(5, 12))

      frame_botones = ctk.CTkFrame(frame_inferior, fg_color="transparent")
      frame_botones.pack()

      # BOTÓN DESCARGAR / ABRIR PDF
      boton_pdf = ctk.CTkButton(
          frame_botones,
          text="Descargar / Abrir PDF",
          command=lambda: self.descargar_pdf(
              id_presupuesto, ventana_detalle
          ),
          fg_color="#1f538d",
          width=160,
          height=35,
      )
      boton_pdf.pack(side="left", padx=10)

      boton_cerrar = ctk.CTkButton(
          frame_botones,
          text="Cerrar",
          command=ventana_detalle.destroy,
          width=120,
          height=35,
      )
      boton_cerrar.pack(side="left", padx=10)

    except Exception as error:
      messagebox.showerror(
          "Error",
          f"No se pudo mostrar el detalle del presupuesto.\n\n{error}",
          parent=self.ventana,
      )

  # ==========================================
  # DESCARGAR / ABRIR PDF
  # ==========================================
  def descargar_pdf(self, id_presupuesto, ventana_padre):
    # Buscar el PDF generado en el sistema
    nombre_archivo = f"presupuesto_{id_presupuesto}.pdf"
    ruta_pdf = self.carpeta_pdf / nombre_archivo

    if ruta_pdf.exists():
      try:
        if sys.platform == "win32":
          os.startfile(ruta_pdf)
        elif sys.platform == "darwin":
          subprocess.call(["open", ruta_pdf])
        else:
          subprocess.call(["xdg-open", ruta_pdf])

        messagebox.showinfo(
            "PDF",
            f"Se abrió correctamente el presupuesto:\n{ruta_pdf}",
            parent=ventana_padre,
        )
      except Exception as e:
        messagebox.showerror(
            "Error",
            f"No se pudo abrir el archivo PDF.\n\n{e}",
            parent=ventana_padre,
        )
    else:
      messagebox.showwarning(
          "PDF No encontrado",
          f"No se encontró el archivo PDF en la ruta:\n{ruta_pdf}\n\nVerifique si fue generado previamente.",
          parent=ventana_padre,
      )