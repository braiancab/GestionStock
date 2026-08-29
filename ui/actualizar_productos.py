from pathlib import Path
import customtkinter as ctk
from openpyxl import load_workbook
from tkinter import messagebox


class VentanaActualizarPrecios:

  def __init__(self, ventana_padre, callback_actualizado=None):
    self.ventana_padre = ventana_padre
    self.callback_actualizado = callback_actualizado

    # Ruta base de los archivos Excel por rubro
    self.carpeta_datos = (
        Path.home()
        / "AppData"
        / "Local"
        / "GestionStock"
        / "datos"
    )

    # Diccionario de rubros y sus archivos correspondientes
    self.rubros_archivos = {
        "Ferretería": "ferreteria.xlsx",
        "Refrigeración": "refrigeracion.xlsx",
        "Electricidad": "electricidad.xlsx",
    }

    # ==========================================
    # VENTANA
    # ==========================================
    self.ventana = ctk.CTkToplevel(ventana_padre)
    self.ventana.title("Actualizar Precios por Rubro")
    self.ventana.geometry("500x420")
    self.ventana.resizable(False, False)

    self.ventana.transient(ventana_padre)
    self.ventana.grab_set()

    self.crear_interfaz()

  # ==========================================
  # INTERFAZ
  # ==========================================
  def crear_interfaz(self):
    # Título
    titulo = ctk.CTkLabel(
        self.ventana,
        text="ACTUALIZAR PRECIOS",
        font=ctk.CTkFont(size=22, weight="bold"),
    )
    titulo.pack(pady=(25, 15))

    # Frame principal de formulario
    frame_form = ctk.CTkFrame(self.ventana)
    frame_form.pack(fill="both", expand=True, padx=30, pady=10)

    # Seleccionar Rubro
    lbl_rubro = ctk.CTkLabel(
        frame_form,
        text="Seleccionar Rubro:",
        font=ctk.CTkFont(size=14, weight="bold"),
    )
    lbl_rubro.pack(anchor="w", padx=20, pady=(15, 5))

    opciones_rubro = ["Todos los rubros"] + list(self.rubros_archivos.keys())
    self.combo_rubro = ctk.CTkOptionMenu(
        frame_form, values=opciones_rubro, width=380, height=35
    )
    self.combo_rubro.pack(padx=20, pady=(0, 15))

    # Porcentaje de Aumento/Descuento
    lbl_porcentaje = ctk.CTkLabel(
        frame_form,
        text="Porcentaje de modificación (%):",
        font=ctk.CTkFont(size=14, weight="bold"),
    )
    lbl_porcentaje.pack(anchor="w", padx=20, pady=(5, 5))

    self.entry_porcentaje = ctk.CTkEntry(
        frame_form,
        width=380,
        height=35,
        placeholder_text="Ej: 15 (aumenta 15%) o -10 (descuento 10%)",
    )
    self.entry_porcentaje.pack(padx=20, pady=(0, 10))

    lbl_info = ctk.CTkLabel(
        frame_form,
        text="* Use números positivos para aumentos y negativos para descuentos.",
        font=ctk.CTkFont(size=11),
        text_color="gray",
    )
    lbl_info.pack(anchor="w", padx=20, pady=(0, 15))

    # Botones
    frame_botones = ctk.CTkFrame(self.ventana, fg_color="transparent")
    frame_botones.pack(fill="x", padx=30, pady=(10, 20))

    btn_aplicar = ctk.CTkButton(
        frame_botones,
        text="Aplicar Cambio",
        command=self.aplicar_actualizacion,
        fg_color="#1f538d",
        hover_color="#14375e",
        width=180,
        height=38,
        font=ctk.CTkFont(size=14, weight="bold"),
    )
    btn_aplicar.pack(side="left", padx=(20, 10))

    btn_cancelar = ctk.CTkButton(
        frame_botones,
        text="Cancelar",
        command=self.ventana.destroy,
        fg_color="gray",
        hover_color="#555555",
        width=140,
        height=38,
        font=ctk.CTkFont(size=14),
    )
    btn_cancelar.pack(side="right", padx=(10, 20))

  # ==========================================
  # LÓGICA DE ACTUALIZACIÓN DE PRECIOS
  # ==========================================
  def aplicar_actualizacion(self):
      texto_porcentaje = self.entry_porcentaje.get().strip().replace(",", ".")

      try:
        porcentaje = float(texto_porcentaje)
        if porcentaje == 0:
          messagebox.showwarning(
              "Atención",
              "El porcentaje ingresado debe ser distinto de 0.",
              parent=self.ventana,
          )
          return
      except ValueError:
        messagebox.showerror(
            "Error",
            "Por favor, ingrese un número válido para el porcentaje.",
            parent=self.ventana,
        )
        return

      rubro_seleccionado = self.combo_rubro.get()

      tipo_operacion = "Aumento" if porcentaje > 0 else "Descuento"
      mensaje_conf = (
          f"¿Está seguro de aplicar un {tipo_operacion} del {abs(porcentaje)}% "
          f"en los productos de: '{rubro_seleccionado}'?"
      )

      confirmar = messagebox.askyesno(
          "Confirmar Actualización", mensaje_conf, parent=self.ventana
      )
      if not confirmar:
        return

      # Determinar qué archivos procesar
      archivos_a_procesar = []
      if rubro_seleccionado == "Todos los rubros":
        archivos_a_procesar = list(self.rubros_archivos.values())
      else:
        archivos_a_procesar = [self.rubros_archivos[rubro_seleccionado]]

      factor_multiplicador = 1 + (porcentaje / 100.0)
      productos_modificados = 0

      try:
        for nombre_archivo in archivos_a_procesar:
          ruta_excel = self.carpeta_datos / nombre_archivo

          if not ruta_excel.exists():
            continue

          # Cargar libro sin data_only para permitir escritura
          wb = load_workbook(ruta_excel)
          ws = wb.active

          # Recorrer las filas directamente desde la fila 2 hasta la última fila con datos
          for fila_idx in range(2, ws.max_row + 1):
            celda_id = ws.cell(row=fila_idx, column=1)  # Columna A (ID)
            celda_precio = ws.cell(
                row=fila_idx, column=6
            )  # Columna F (Precio - Columna 6)

            # Si la fila tiene un ID de producto válido
            if celda_id.value is not None:
              val_precio = celda_precio.value

              if val_precio is not None:
                # Limpiar posibles caracteres de texto
                val_clean = (
                    str(val_precio)
                    .replace("$", "")
                    .replace(",", "")
                    .replace(" ", "")
                    .strip()
                )

                try:
                  precio_actual = float(val_clean)
                  nuevo_precio = round(precio_actual * factor_multiplicador, 2)

                  # Guardar el número resultante
                  celda_precio.value = nuevo_precio
                  productos_modificados += 1
                except ValueError:
                  continue

          wb.save(ruta_excel)
          wb.close()

        messagebox.showinfo(
            "Éxito",
            f"Se actualizaron correctamente {productos_modificados} productos.",
            parent=self.ventana,
        )

        # Refrescar la tabla en la ventana principal
        if self.callback_actualizado:
          self.callback_actualizado()

        self.ventana.destroy()

      except Exception as error:
        messagebox.showerror(
            "Error",
            f"Ocurrió un error al intentar actualizar los precios:\n\n{error}",
            parent=self.ventana,
        )