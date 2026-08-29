from datetime import datetime
import os
from pathlib import Path

# Si usas ReportLab para generar los archivos PDF:
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from openpyxl import Workbook, load_workbook


class PresupuestoService:

  def __init__(self):
    self.carpeta_datos = (
        Path.home()
        / "AppData"
        / "Local"
        / "GestionStock"
        / "datos"
    )

    self.carpeta_datos.mkdir(parents=True, exist_ok=True)

    self.ruta_excel = self.carpeta_datos / "presupuestos.xlsx"

    # ==========================================
    # CARPETA DE PDFs
    # ==========================================
    self.carpeta_pdf = (
        Path.home()
        / "AppData"
        / "Local"
        / "GestionStock"
        / "presupuestos_pdf"
    )

    # Crea la carpeta si no existe
    self.carpeta_pdf.mkdir(parents=True, exist_ok=True)

    self.encabezados = [
        "ID Presupuesto",
        "Fecha",
        "Cliente",
        "ID Producto",
        "Producto",
        "Descripcion",
        "Rubro",
        "Cantidad",
        "Precio Unitario",
        "Subtotal",
        "Total Presupuesto",
        "Estado",
    ]

    self.crear_excel_si_no_existe()

  # ==========================================
  # CREAR EXCEL
  # ==========================================
  def crear_excel_si_no_existe(self):
    if not self.ruta_excel.exists():
      wb = Workbook()
      ws = wb.active
      ws.title = "Presupuestos"
      ws.append(self.encabezados)
      wb.save(self.ruta_excel)
      wb.close()

  # ==========================================
  # OBTENER SIGUIENTE ID
  # ==========================================
  def obtener_siguiente_id(self):
    wb = load_workbook(self.ruta_excel)
    ws = wb["Presupuestos"]

    ids = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
      if isinstance(fila[0], int):
        ids.append(fila[0])

    wb.close()

    if not ids:
      return 1

    return max(ids) + 1

  # ==========================================
  # REGISTRAR PRESUPUESTO
  # ==========================================
  def registrar_presupuesto(self, cliente, productos, total_presupuesto):
    wb = load_workbook(self.ruta_excel)
    ws = wb["Presupuestos"]

    id_presupuesto = self.obtener_siguiente_id()
    fecha = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    for producto in productos:
      subtotal = producto["cantidad"] * producto["precio"]

      ws.append([
          id_presupuesto,
          fecha,
          cliente,
          producto["id"],
          producto["nombre"],
          producto["descripcion"],
          producto["rubro"],
          producto["cantidad"],
          producto["precio"],
          subtotal,
          total_presupuesto,
          "Pendiente",
      ])

    wb.save(self.ruta_excel)
    wb.close()

    # GENERAR PDF AUTOMÁTICAMENTE EN LA CARPETA
    self.generar_pdf(id_presupuesto, fecha, cliente, productos, total_presupuesto)

    return id_presupuesto

  # ==========================================
  # GENERAR PDF DE PRESUPUESTO
  # ==========================================
  def generar_pdf(
    self, id_presupuesto, fecha, cliente, productos, total_presupuesto
    ):
    ruta_archivo = self.carpeta_pdf / f"presupuesto_{id_presupuesto}.pdf"

    doc = SimpleDocTemplate(
        str(ruta_archivo),
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40,
    )

    estilos = getSampleStyleSheet()

    # Estilos personalizados alineados con la plantilla
    estilo_empresa = ParagraphStyle(
        "EmpresaTitle",
        parent=estilos["Normal"],
        fontName="Helvetica-Bold",
        fontSize=28,
        leading=32,
        alignment=1,  # Centrado
    )

    estilo_subempresa = ParagraphStyle(
        "EmpresaSub",
        parent=estilos["Normal"],
        fontName="Helvetica",
        fontSize=10,
        leading=14,
        alignment=1,  # Centrado
    )

    estilo_titulo_doc = ParagraphStyle(
        "DocTitle",
        parent=estilos["Normal"],
        fontName="Helvetica-Bold",
        fontSize=20,
        leading=24,
        alignment=1,  # Centrado
    )

    estilo_celda = ParagraphStyle(
        "CeldaNorm",
        parent=estilos["Normal"],
        fontName="Helvetica",
        fontSize=10,
        leading=12,
    )

    estilo_celda_bold = ParagraphStyle(
        "CeldaBold",
        parent=estilos["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=12,
    )

    estilo_total = ParagraphStyle(
        "TotalText",
        parent=estilos["Normal"],
        fontName="Helvetica-Bold",
        fontSize=14,
        leading=18,
        alignment=2,  # Derecha
    )

    elementos = []

    # 1. Cabecera
    elementos.append(Paragraph("SAN PABLO", estilo_empresa))
    elementos.append(Spacer(1, 8))
    elementos.append(
        Paragraph(
            "Teléfono: 3756-418521<br/>Dirección: Lisandro de la Torre & Manuel"
            " Ledesma,<br/>Gobernador Ingeniero Valentín Virasoro",
            estilo_subempresa,
        )
    )
    elementos.append(Spacer(1, 25))

    # 2. Título de Presupuesto
    elementos.append(Paragraph("PRESUPUESTO", estilo_titulo_doc))
    elementos.append(Spacer(1, 20))

    # 3. Tabla de Metadatos (Presupuesto Nº, Fecha, Cliente)
    datos_meta = [
        [
            Paragraph("Presupuesto Nº:", estilo_celda_bold),
            Paragraph(str(id_presupuesto), estilo_celda),
        ],
        [
            Paragraph("Fecha:", estilo_celda_bold),
            Paragraph(str(fecha), estilo_celda),
        ],
        [
            Paragraph("Cliente:", estilo_celda_bold),
            Paragraph(str(cliente), estilo_celda),
        ],
    ]

    tabla_meta = Table(datos_meta, colWidths=[130, 370])
    tabla_meta.setStyle(
        TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ])
    )
    elementos.append(tabla_meta)
    elementos.append(Spacer(1, 25))

    # 4. Tabla de Productos
    # Encabezados en azul oscuro/acero (#2D709C) y texto blanco
    color_azul_cabecera = colors.HexColor("#2D709C")

    datos_productos = [
        [
            Paragraph("<b>Producto</b>", estilos["Normal"]),
            Paragraph("<b>Descripción</b>", estilos["Normal"]),
            Paragraph("<b>Cant.</b>", estilos["Normal"]),
            Paragraph("<b>Precio</b>", estilos["Normal"]),
            Paragraph("<b>Subtotal</b>", estilos["Normal"]),
        ]
    ]

    for prod in productos:
        subtotal = prod["cantidad"] * prod["precio"]
        datos_productos.append([
            Paragraph(str(prod["nombre"]), estilo_celda),
            Paragraph(str(prod.get("descripcion", "")), estilo_celda),
            str(prod["cantidad"]),
            f"${prod['precio']:,.2f}",
            f"${subtotal:,.2f}",
        ])

    tabla_prod = Table(
        datos_productos, colWidths=[120, 180, 50, 75, 75], hAlign="CENTER"
    )
    tabla_prod.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), color_azul_cabecera),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("ALIGN", (2, 0), (2, -1), "CENTER"),  # Cantidad centrada
            ("ALIGN", (3, 0), (-1, -1), "RIGHT"),  # Precio y Subtotal a la derecha
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ])
    )
    elementos.append(tabla_prod)
    elementos.append(Spacer(1, 30))

    # 5. Total General (Alineado a la derecha)
    elementos.append(
        Paragraph(f"TOTAL: ${total_presupuesto:,.2f}", estilo_total)
    )

    doc.build(elementos)
    return ruta_archivo