from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook


class PresupuestoService:

    def __init__(self):

        self.carpeta_datos = (
            Path.home()
            / "AppData"
            / "Local"
            / "GestionStock"
            / "datos"
        )

        self.carpeta_datos.mkdir(
            parents=True,
            exist_ok=True
        )

        self.ruta_excel = (
            self.carpeta_datos
            / "presupuestos.xlsx"
        )

        self.encabezados = [
            "ID Presupuesto",
            "Fecha",
            "Cliente",
            "Telefono",
            "ID Producto",
            "Producto",
            "Descripcion",
            "Rubro",
            "Cantidad",
            "Precio Unitario",
            "Subtotal",
            "Total Presupuesto",
            "Estado"
        ]

        self.crear_excel_si_no_existe()



    # ==========================================
    # CREAR EXCEL
    # ==========================================

    def crear_excel_si_no_existe(self):

        if not self.ruta_excel.exists():

            wb = Workbook()

            ws = wb.active

            ws.title = "Presupuestos"

            ws.append(
                self.encabezados
            )

            wb.save(
                self.ruta_excel
            )

            wb.close()

        # ==========================================
    # SIGUIENTE ID
    # ==========================================

    def obtener_siguiente_id(self):

        wb = load_workbook(
            self.ruta_excel
        )

        ws = wb["Presupuestos"]

        ids = []

        for fila in ws.iter_rows(
            min_row=2,
            values_only=True
        ):

            id_presupuesto = fila[0]

            if isinstance(
                id_presupuesto,
                int
            ):
                ids.append(
                    id_presupuesto
                )

        wb.close()

        if not ids:
            return 1

        return max(ids) + 1


        # ==========================================
    # REGISTRAR PRESUPUESTO
    # ==========================================

    def registrar_presupuesto(
        self,
        cliente,
        telefono,
        productos,
        total_presupuesto
    ):

        wb = load_workbook(
            self.ruta_excel
        )

        ws = wb["Presupuestos"]

        id_presupuesto = (
            self.obtener_siguiente_id()
        )

        fecha = datetime.now().strftime(
            "%d/%m/%Y %H:%M:%S"
        )

        for producto in productos:

            subtotal = (
                producto["cantidad"]
                * producto["precio"]
            )

            ws.append([

                id_presupuesto,

                fecha,

                cliente,

                telefono,

                producto["id"],

                producto["nombre"],

                producto.get(
                    "descripcion",
                    ""
                ),

                producto.get(
                    "rubro",
                    ""
                ),

                producto["cantidad"],

                producto["precio"],

                subtotal,

                total_presupuesto,

                "Pendiente"

            ])

        wb.save(
            self.ruta_excel
        )

        wb.close()

        return id_presupuesto

        # ==========================================
    # OBTENER PRESUPUESTOS
    # ==========================================

    def obtener_presupuestos(self):

        wb = load_workbook(
            self.ruta_excel,
            data_only=True
        )

        ws = wb["Presupuestos"]

        presupuestos = {}

        for fila in ws.iter_rows(
            min_row=2,
            values_only=True
        ):

            if fila[0] is None:
                continue

            id_presupuesto = fila[0]

            if id_presupuesto not in presupuestos:

                presupuestos[
                    id_presupuesto
                ] = {

                    "id": id_presupuesto,

                    "fecha": fila[1],

                    "cliente": fila[2],

                    

                    "total": fila[11],

                    "estado": fila[12],

                    "productos": []

                }

            producto = {

                "id": fila[4],

                "nombre": fila[5],

                "descripcion": fila[6],

                "rubro": fila[7],

                "cantidad": fila[8],

                "precio": fila[9],

                "subtotal": fila[10]

            }

            presupuestos[
                id_presupuesto
            ]["productos"].append(
                producto
            )

        wb.close()

        return list(
            presupuestos.values()
        )