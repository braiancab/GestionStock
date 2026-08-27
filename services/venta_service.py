from pathlib import Path
from openpyxl import Workbook, load_workbook
from datetime import datetime


class VentaService:

    def __init__(self):

        # ==========================================
        # CARPETA DE DATOS DE GESTIONSTOCK
        # ==========================================

        carpeta_datos = (
            Path.home()
            / "AppData"
            / "Local"
            / "GestionStock"
            / "datos"
        )

        # Crear la carpeta si no existe
        carpeta_datos.mkdir(
            parents=True,
            exist_ok=True
        )

        # Archivo de ventas
        self.ruta_excel = (
            carpeta_datos
            / "ventas.xlsx"
        )

        self.encabezados = [
            "ID Venta",
            "Fecha",
            "Cliente",
            "ID Producto",
            "Producto",
            "Variante",
            "Cantidad",
            "Precio Unitario",
            "Subtotal",
            "Total Venta"
        ]

        self.crear_excel_si_no_existe()


    # ==========================================
    # CREAR EXCEL DE VENTAS
    # ==========================================

    def crear_excel_si_no_existe(self):

        if not self.ruta_excel.exists():

            wb = Workbook()

            ws = wb.active

            ws.title = "Ventas"

            ws.append(self.encabezados)

            wb.save(self.ruta_excel)


    # ==========================================
    # OBTENER SIGUIENTE ID DE VENTA
    # ==========================================

    def obtener_siguiente_id(self):

        wb = load_workbook(
            self.ruta_excel,
            read_only=True
        )

        ws = wb["Ventas"]

        ids = []

        for fila in ws.iter_rows(
            min_row=2,
            values_only=True
        ):

            id_venta = fila[0]

            if isinstance(id_venta, int):

                ids.append(id_venta)

        wb.close()

        if not ids:

            return 1

        return max(ids) + 1


    # ==========================================
    # REGISTRAR VENTA
    # ==========================================

    def registrar_venta(
        self,
        cliente,
        productos,
        total_venta
    ):

        wb = load_workbook(
            self.ruta_excel
        )

        ws = wb["Ventas"]

        id_venta = self.obtener_siguiente_id()

        fecha = datetime.now().strftime(
            "%d/%m/%Y %H:%M:%S"
        )

        for producto in productos:

            subtotal = (
                producto["cantidad"]
                * producto["precio"]
            )

            ws.append([
                id_venta,
                fecha,
                cliente,
                producto["id"],
                producto["nombre"],
                producto["descripcion"],
                producto["cantidad"],
                producto["precio"],
                subtotal,
                total_venta
            ])

        wb.save(
            self.ruta_excel
        )

        wb.close()

        return id_venta