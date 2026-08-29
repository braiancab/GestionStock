from pathlib import Path
from openpyxl import Workbook, load_workbook


class ExcelService:

    RUBROS = {
        "ferreteria": "stock_tienda_ferreteria.xlsx",
        "refrigeracion": "stock_tienda_refrigeracion.xlsx",
        "electricidad": "stock_tienda_electricidad.xlsx"
    }

    def __init__(self, rubro=None):

        # Carpeta de datos de GestionStock
        self.carpeta_datos = (
            Path.home()
            / "AppData"
            / "Local"
            / "GestionStock"
            / "datos"
        )

        self.carpeta_datos.mkdir(
            parents=True,
            exist_ok=True
        )

        self.rubro = rubro

        # Si se especificó un rubro,
        # utilizar su Excel correspondiente.
        if rubro in self.RUBROS:

            self.ruta_excel = (
                self.carpeta_datos / self.RUBROS[rubro]
            )

        else:

            # Ruta utilizada para operaciones
            # que necesitan trabajar con los 3 rubros.
            self.ruta_excel = None

        self.encabezados = [
            "ID",
            "Nombre",
            "Descripcion",
            "Categoria",
            "Stock",
            "Precio"
        ]

        # Crear Excel del rubro si corresponde
        if self.ruta_excel is not None:

            self.crear_excel_si_no_existe()


    # ==========================================
    # CREAR EXCEL
    # ==========================================

    def crear_excel_si_no_existe(self):

        if not self.ruta_excel.exists():

            wb = Workbook()

            ws = wb.active

            ws.title = "Productos"

            ws.append(self.encabezados)

            wb.save(self.ruta_excel)

            wb.close()

            print(
                f"Excel creado correctamente: "
                f"{self.ruta_excel}"
            )


    # ==========================================
    # OBTENER PRODUCTOS
    # ==========================================

    def obtener_productos(self):

        if self.ruta_excel is None:

            raise ValueError(
                "No se especificó un rubro."
            )

        wb = load_workbook(
            self.ruta_excel
        )

        ws = wb["Productos"]

        productos = []

        for fila in ws.iter_rows(
            min_row=2,
            values_only=True
        ):

            if fila[0] is None:
                continue

            producto = {

                "id": fila[0],

                "nombre": fila[1],

                "descripcion": fila[2],

                "categoria": fila[3],

                "stock": fila[4] or 0,

                "precio": fila[5] or 0
            }

            productos.append(producto)

        wb.close()

        return productos


    # ==========================================
    # OBTENER PRODUCTOS DE LOS 3 RUBROS
    # ==========================================

    def obtener_productos_todos_los_rubros(self):

        productos_totales = []

        for rubro in self.RUBROS:

            servicio = ExcelService(rubro)

            productos = servicio.obtener_productos()

            for producto in productos:

                producto["rubro"] = rubro

            productos_totales.extend(
                productos
            )

        return productos_totales


    # ==========================================
    # OBTENER SIGUIENTE ID
    # ==========================================

    def obtener_siguiente_id(self):

        productos = self.obtener_productos()

        if not productos:

            return 1

        ids = [

            producto["id"]

            for producto in productos

            if isinstance(
                producto["id"],
                int
            )
        ]

        if not ids:

            return 1

        return max(ids) + 1


    # ==========================================
    # AGREGAR PRODUCTO
    # ==========================================

    def agregar_producto(
        self,
        nombre,
        descripcion,
        categoria,
        stock,
        precio
    ):

        if self.ruta_excel is None:

            raise ValueError(
                "No se especificó un rubro."
            )

        wb = load_workbook(
            self.ruta_excel
        )

        ws = wb["Productos"]

        nuevo_id = (
            self.obtener_siguiente_id()
        )

        nueva_fila = [

            nuevo_id,

            nombre,

            descripcion,

            categoria,

            stock,

            precio
        ]

        ws.append(
            nueva_fila
        )

        wb.save(
            self.ruta_excel
        )

        wb.close()

        print(
            f"Producto agregado correctamente. "
            f"ID: {nuevo_id}"
        )

        return nuevo_id


    # ==========================================
    # ACTUALIZAR PRODUCTO
    # ==========================================

    def actualizar_producto(
        self,
        id_producto,
        nombre,
        descripcion,
        categoria,
        stock,
        precio
    ):

        if self.ruta_excel is None:

            raise ValueError(
                "No se especificó un rubro."
            )

        wb = load_workbook(
            self.ruta_excel
        )

        ws = wb["Productos"]

        producto_encontrado = False

        for fila in range(
            2,
            ws.max_row + 1
        ):

            id_actual = ws.cell(
                row=fila,
                column=1
            ).value

            if id_actual == id_producto:

                ws.cell(
                    row=fila,
                    column=2
                ).value = nombre

                ws.cell(
                    row=fila,
                    column=3
                ).value = descripcion

                ws.cell(
                    row=fila,
                    column=4
                ).value = categoria

                ws.cell(
                    row=fila,
                    column=5
                ).value = stock

                ws.cell(
                    row=fila,
                    column=6
                ).value = precio

                producto_encontrado = True

                break

        if not producto_encontrado:

            wb.close()

            raise ValueError(
                f"No se encontró el producto "
                f"con ID {id_producto}"
            )

        wb.save(
            self.ruta_excel
        )

        wb.close()

        print(
            f"Producto con ID {id_producto} "
            f"actualizado correctamente."
        )


    # ==========================================
    # ELIMINAR PRODUCTO
    # ==========================================

    def eliminar_producto(
        self,
        id_producto
    ):

        if self.ruta_excel is None:

            raise ValueError(
                "No se especificó un rubro."
            )

        wb = load_workbook(
            self.ruta_excel
        )

        ws = wb["Productos"]

        producto_encontrado = False

        for fila in range(
            2,
            ws.max_row + 1
        ):

            id_actual = ws.cell(
                row=fila,
                column=1
            ).value

            if id_actual == id_producto:

                ws.delete_rows(
                    fila,
                    1
                )

                producto_encontrado = True

                break

        if not producto_encontrado:

            wb.close()

            raise ValueError(
                f"No se encontró el producto "
                f"con ID {id_producto}"
            )

        wb.save(
            self.ruta_excel
        )

        wb.close()

        print(
            f"Producto con ID {id_producto} "
            f"eliminado correctamente."
        )


    # ==========================================
    # DESCONTAR STOCK
    # ==========================================

    def descontar_stock(
        self,
        id_producto,
        cantidad,
        rubro=None
    ):

        # Si se especifica un rubro,
        # utilizar ese Excel.
        if rubro:

            servicio = ExcelService(
                rubro
            )

            servicio.descontar_stock(
                id_producto,
                cantidad
            )

            return

        if self.ruta_excel is None:

            raise ValueError(
                "No se especificó un rubro."
            )

        wb = load_workbook(
            self.ruta_excel
        )

        ws = wb["Productos"]

        producto_encontrado = False

        for fila in range(
            2,
            ws.max_row + 1
        ):

            id_actual = ws.cell(
                row=fila,
                column=1
            ).value

            if id_actual == id_producto:

                stock_actual = ws.cell(
                    row=fila,
                    column=5
                ).value

                if stock_actual is None:

                    stock_actual = 0

                if cantidad > stock_actual:

                    wb.close()

                    raise ValueError(
                        f"No hay suficiente stock.\n\n"
                        f"Stock disponible: "
                        f"{stock_actual}\n"
                        f"Cantidad solicitada: "
                        f"{cantidad}"
                    )

                nuevo_stock = (
                    stock_actual - cantidad
                )

                ws.cell(
                    row=fila,
                    column=5
                ).value = nuevo_stock

                producto_encontrado = True

                break

        if not producto_encontrado:

            wb.close()

            raise ValueError(
                f"No se encontró el producto "
                f"con ID {id_producto}"
            )

        wb.save(
            self.ruta_excel
        )

        wb.close()


    # ==========================================
    # OBTENER VARIANTES
    # ==========================================

    def obtener_variantes(
        self,
        nombre_producto
    ):

        productos = (
            self.obtener_productos()
        )

        variantes = []

        for producto in productos:

            nombre = str(
                producto["nombre"] or ""
            ).strip()

            if (
                nombre.lower()
                ==
                nombre_producto.lower()
            ):

                variantes.append(
                    producto
                )

        return variantes


    # ==========================================
    # ACTUALIZAR PRECIOS POR PORCENTAJE
    # ==========================================
    # ==========================================
    # ACTUALIZAR PRECIOS POR PORCENTAJE
    # ==========================================
    def actualizar_precios_por_porcentaje(self, porcentaje, rubro=None):
      """Aplica un porcentaje de aumento o descuento a los precios.

      :param porcentaje: float (ej. 15 para +15%, -10 para -10%)
      :param rubro: str ('ferreteria', 'refrigeracion', 'electricidad' o None
      para todos)
      :return: int (cantidad de productos modificados)
      """
      factor = 1 + (porcentaje / 100.0)
      productos_modificados = 0

      # Determinar qué rubros procesar
      rubros_a_procesar = [rubro] if rubro in self.RUBROS else list(self.RUBROS)

      for r in rubros_a_procesar:
        ruta = self.carpeta_datos / self.RUBROS[r]

        if not ruta.exists():
          continue

        wb = load_workbook(ruta)
        ws = wb["Productos"]

        for fila in range(2, ws.max_row + 1):
          id_celda = ws.cell(row=fila, column=1).value
          precio_celda = ws.cell(row=fila, column=6).value

          if id_celda is not None and precio_celda is not None:
            try:
              val_clean = (
                  str(precio_celda)
                  .replace("$", "")
                  .replace(",", "")
                  .replace(" ", "")
                  .strip()
              )
              precio_actual = float(val_clean)
              nuevo_precio = round(precio_actual * factor, 2)

              # Guardar el nuevo valor
              ws.cell(row=fila, column=6).value = nuevo_precio
              productos_modificados += 1
            except ValueError:
              continue

        wb.save(ruta)
        wb.close()

      return productos_modificados